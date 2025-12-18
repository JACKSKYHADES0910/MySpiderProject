# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook('output/HK004_Projects_20251215_135531.xlsx')
ws = wb.active

print("=== 项目名称示例 (前5行) ===")
for i in range(2, min(7, ws.max_row + 1)):  # 跳过表头
    项目名称 = ws.cell(row=i, column=3).value  # 第3列是项目名称
    print(f"{i-1}. {项目名称}")

print(f"\n总行数（含表头）: {ws.max_row}")
print(f"项目总数: {ws.max_row - 1}")
