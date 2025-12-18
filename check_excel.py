# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook('output/HK004_Projects_20251215_134409.xlsx')
ws = wb.active

print("=== 前5行数据 ===")
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    print(f"Row {i}: {row}")
    if i >= 5:
        break

print(f"\n总行数: {ws.max_row}")
print(f"总列数: {ws.max_column}")
